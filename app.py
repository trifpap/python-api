from flask import Flask, request, send_file, jsonify
import pandas as pd
import io

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

app = Flask(__name__)

@app.route("/")
def home():
    return "Excel API is running!"

#@app.route("/process-excel", methods=["POST"])
#def process_excel():

#    if 'file' not in request.files:
#        return jsonify({"error": "No file uploaded"}), 400

#    file = request.files['file']

#    try:
#        df = pd.read_excel(file)

#        # Example processing 
        
        # Remove completely empty rows
#        df.dropna(how='all', inplace=True)     

        # Clean column names
#        df.columns = [col.strip().upper() for col in df.columns]

        # Remove pandas duplicate suffix (.1, .2 etc.)
#        df.columns = df.columns.str.replace(r'\.\d+$', '', regex=True)

        # Remove duplicate columns (keep first occurrence)
#        df = df.loc[:, ~df.columns.duplicated()]

        # Remove duplicate rows
#        df = df.drop_duplicates()

#        output = io.BytesIO()
#        df.to_excel(output, index=False)
#        output.seek(0)

#        return send_file(
#            output,
#            download_name="processed.xlsx",
#            as_attachment=True,
#            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#        )

#    except Exception as e:
#        return jsonify({"error": str(e)}), 500
    

@app.route("/process-excel", methods=["POST"])
def process_excel():

    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']

    try:
        df = pd.read_excel(file)

        # ---------------- CLEANING ----------------
        df.dropna(how='all', inplace=True)
        df.columns = [col.strip().upper() for col in df.columns]
        df.columns = df.columns.str.replace(r'\.\d+$', '', regex=True)
        df = df.loc[:, ~df.columns.duplicated()]

        duplicate_rows = df.duplicated().sum()
        df = df.drop_duplicates()

        # ---------------- BASIC METRICS ----------------
        num_rows = len(df)
        num_columns = len(df.columns)
        column_names = ", ".join(df.columns)

        null_counts = df.isnull().sum()
        numeric_df = df.select_dtypes(include='number')

        # ---------------- NUMERIC STATS ----------------
        stats_df = pd.DataFrame()

        if not numeric_df.empty:
            stats_df = pd.DataFrame({
                "Mean": numeric_df.mean(),
                "Median": numeric_df.median(),
                "Std Dev": numeric_df.std(),
                "Min": numeric_df.min(),
                "Max": numeric_df.max()
            })

        # ---------------- COUNTRY FREQUENCY ----------------
        country_freq = pd.DataFrame()
        if "COUNTRY" in df.columns:
            country_freq = df["COUNTRY"].value_counts().reset_index()
            country_freq.columns = ["Country", "Count"]

        # ---------------- DATA QUALITY SCORE ----------------
        total_cells = df.size
        total_nulls = df.isnull().sum().sum()
        quality_score = round((1 - (total_nulls / total_cells)) * 100, 2) if total_cells > 0 else 100

        summary_df = pd.DataFrame({
            "Metric": [
                "Number of Rows",
                "Number of Columns",
                "Duplicate Rows Removed",
                "Data Quality Score (%)",
                "Column Names"
            ],
            "Value": [
                num_rows,
                num_columns,
                duplicate_rows,
                quality_score,
                column_names
            ]
        })

        null_df = null_counts.reset_index()
        null_df.columns = ["Column", "Null Count"]

        # ---------------- WRITE TO EXCEL ----------------
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="DATA", index=False)
            summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
            stats_df.to_excel(writer, sheet_name="NUMERIC_STATS")
            null_df.to_excel(writer, sheet_name="NULL_COUNTS", index=False)

            if not country_freq.empty:
                country_freq.to_excel(writer, sheet_name="COUNTRY_FREQ", index=False)

            workbook = writer.book

            # ---------------- MEAN BAR CHART ----------------
            if not stats_df.empty:
                sheet = writer.sheets["NUMERIC_STATS"]
                chart = BarChart()
                chart.title = "Mean Values"
                chart.y_axis.title = "Mean"
                chart.x_axis.title = "Columns"

                data = Reference(sheet, min_col=2, min_row=1,
                                 max_col=2, max_row=len(stats_df)+1)
                cats = Reference(sheet, min_col=1, min_row=2,
                                 max_row=len(stats_df)+1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                sheet.add_chart(chart, "H2")

            # ---------------- COUNTRY BAR CHART ----------------
            if not country_freq.empty:
                sheet = writer.sheets["COUNTRY_FREQ"]
                chart = BarChart()
                chart.title = "Country Distribution"
                chart.y_axis.title = "Count"

                data = Reference(sheet, min_col=2, min_row=1,
                                 max_col=2, max_row=len(country_freq)+1)
                cats = Reference(sheet, min_col=1, min_row=2,
                                 max_row=len(country_freq)+1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                sheet.add_chart(chart, "E2")

            # ---------------- DATA QUALITY VISUAL ----------------
            summary_sheet = writer.sheets["SUMMARY"]
            if quality_score >= 80:
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif quality_score >= 50:
                fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

            summary_sheet["B4"].fill = fill  # Quality score cell

            # ---------------- NULL HEATMAP ----------------
            heatmap_sheet = workbook.create_sheet("NULL_HEATMAP")

            for r in range(len(df)):
                for c in range(len(df.columns)):
                    cell = heatmap_sheet.cell(row=r+1, column=c+1)
                    if pd.isnull(df.iloc[r, c]):
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        output.seek(0)

        return send_file(
            output,
            download_name="advanced_excel_report.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500